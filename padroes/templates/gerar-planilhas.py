"""
Gera as planilhas modelo da Borin, com a identidade aplicada.

Construidas do zero a partir da estrutura funcional descrita em
`padroes/padrao-entrega.md`. Nao contem arquivo, macro, aba ou base de dados
de terceiros — so a estrutura, que e boa pratica de engenharia.

    python padroes/templates/gerar-planilhas.py

Saida: padroes/templates/xlsx/
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

AQUI = os.path.dirname(os.path.abspath(__file__))
SAIDA = os.path.join(AQUI, "xlsx")

# ------------------------------------------------------------------ identidade

TINTA = "111111"
CINZA_600 = "6B6B6B"
CINZA_300 = "DCDCDC"
CINZA_100 = "F4F4F4"
COMANDO = "C1121F"

SANS = "Inter"
MONO = "Consolas"

f_marca = Font(name=SANS, size=16, bold=True, color=TINTA)
f_campo = Font(name=MONO, size=7, color=CINZA_600)
f_valor = Font(name=MONO, size=10, color=TINTA)
f_cabec = Font(name=SANS, size=10, bold=True, color=TINTA)
f_corpo = Font(name=SANS, size=10, color=TINTA)
f_mono = Font(name=MONO, size=10, color=TINTA)
f_nota = Font(name=SANS, size=9, italic=True, color=CINZA_600)
f_secao = Font(name=SANS, size=11, bold=True, color=TINTA)

b_cabec = Border(bottom=Side(style="medium", color=TINTA))
b_linha = Border(bottom=Side(style="thin", color=CINZA_300))
b_topo = Border(top=Side(style="thin", color=CINZA_300))

fill_zona = PatternFill("solid", fgColor=CINZA_100)


def carimbo(ws, largura_tabela):
    """Bloco de identificacao no topo — o carimbo da marca em planilha."""
    ws["A1"] = "BORIN"
    ws["A1"].font = f_marca
    ws["A2"] = "projetos elétricos industriais"
    ws["A2"].font = Font(name=SANS, size=8, color=CINZA_600)

    campos = ["PROJETO", "CLIENTE", "CÓDIGO", "REV", "DATA"]
    col = 3
    for campo in campos:
        c = ws.cell(row=1, column=col, value=campo)
        c.font = f_campo
        v = ws.cell(row=2, column=col, value="")
        v.font = f_valor
        v.border = b_linha
        col += 1

    for col in range(1, max(largura_tabela, 8) + 1):
        ws.cell(row=3, column=col).border = b_topo
    ws.row_dimensions[3].height = 6


def tabela(ws, linha, cabecalhos, larguras, exemplos=None, mono_cols=()):
    """Escreve o cabecalho da tabela e linhas de exemplo."""
    for i, (titulo, larg) in enumerate(zip(cabecalhos, larguras), start=1):
        c = ws.cell(row=linha, column=i, value=titulo)
        c.font = f_cabec
        c.border = b_cabec
        c.alignment = Alignment(vertical="bottom", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = larg

    if exemplos:
        for j, linha_ex in enumerate(exemplos, start=1):
            for i, valor in enumerate(linha_ex, start=1):
                c = ws.cell(row=linha + j, column=i, value=valor)
                c.font = f_mono if i in mono_cols else f_corpo
                c.border = b_linha
    ws.freeze_panes = ws.cell(row=linha + 1, column=1)
    return linha + (len(exemplos) if exemplos else 0) + 1


def nota(ws, linha, texto):
    c = ws.cell(row=linha, column=1, value=texto)
    c.font = f_nota
    return linha + 1


def novo(titulo_aba):
    wb = Workbook()
    wb.active.title = titulo_aba
    return wb, wb.active


def salvar(wb, nome):
    os.makedirs(SAIDA, exist_ok=True)
    caminho = os.path.join(SAIDA, nome)
    wb.save(caminho)
    print(f"  {nome}")


# ------------------------------------------------------------------ 1. material

def lista_de_materiais():
    wb, ws = novo("Lista de Materiais")
    carimbo(ws, 7)
    fim = tabela(
        ws, 5,
        ["Instalação", "Código", "Descrição", "Fabricante", "Unidade", "Qtd", "Observação"],
        [16, 14, 46, 18, 10, 8, 30],
        exemplos=[
            ["Painel 1", "", "", "", "un", "", ""],
            ["Painel 1", "", "", "", "m", "", ""],
            ["Campo", "", "", "", "un", "", ""],
        ],
        mono_cols=(2, 6),
    )
    nota(ws, fim + 1, "Uma linha por item. Agrupar por instalação — a lista do PDF sai daqui e as duas têm que bater.")
    salvar(wb, "MODELO - Lista de Materiais.xlsx")


# ------------------------------------------------------------------ 2. CLP

def arquitetura_clp():
    wb, ws = novo("Arquitetura de CLP")
    carimbo(ws, 8)
    fim = tabela(
        ws, 5,
        ["Instalação", "Tag", "Módulo", "Canal", "Endereço", "Tipo", "Função", "Observação"],
        [14, 12, 18, 10, 14, 14, 40, 26],
        exemplos=[
            ["Painel 1", "A1", "CPU", "", "", "", "", ""],
            ["Painel 1", "A2", "", "1", "", "ED", "", ""],
            ["Painel 1", "A2", "", "2", "", "ED", "", ""],
            ["Painel 1", "A3", "", "1", "", "SD", "", ""],
            ["Painel 1", "A4", "", "1", "", "EA", "", ""],
        ],
        mono_cols=(2, 4, 5, 6),
    )

    dv = DataValidation(type="list", formula1='"ED,SD,EA,SA,COM,RESERVA"', allow_blank=True)
    dv.prompt = "ED entrada digital · SD saída digital · EA entrada analógica · SA saída analógica"
    ws.add_data_validation(dv)
    dv.add(f"F6:F500")

    fim = nota(ws, fim + 1, "Todo endereço que aparece no diagrama tem que existir aqui, e sem duplicar.")
    nota(ws, fim, "Declarar a reserva de I/O como linha, com tipo RESERVA — reserva não declarada é reserva que some.")

    ws2 = wb.create_sheet("Resumo")
    ws2["A1"] = "Resumo de pontos"
    ws2["A1"].font = f_secao
    tabela(ws2, 3, ["Tipo", "Usados", "Reserva", "Total"], [22, 12, 12, 12],
           exemplos=[
               ["Entradas digitais", "=COUNTIFS('Arquitetura de CLP'!F:F,\"ED\")", "", ""],
               ["Saídas digitais", "=COUNTIFS('Arquitetura de CLP'!F:F,\"SD\")", "", ""],
               ["Entradas analógicas", "=COUNTIFS('Arquitetura de CLP'!F:F,\"EA\")", "", ""],
               ["Saídas analógicas", "=COUNTIFS('Arquitetura de CLP'!F:F,\"SA\")", "", ""],
           ], mono_cols=(2, 3, 4))
    salvar(wb, "MODELO - Arquitetura de CLP.xlsx")


# ------------------------------------------------------------------ 3. térmico

# Base de dissipação — SÓ ENTRA AQUI O QUE VEIO DE CATÁLOGO, com documento e
# data. Regra #2 do AGENTS.md, valendo pra número técnico e não só pra preço:
# dissipação chutada faz o painel ferver no cliente, e o defeito só aparece no
# primeiro dia quente de janeiro, meses depois da entrega.
#
# Levantado em 05/08/2026 destes documentos:
#   MDWH — WEG, "MDW Minidisjuntores", 50163906, pág. 17, tabela "Dissipação de
#          potência MDWH (norma NBR NM 60898)"
#          static.weg.net/medias/downloadcenter/ha8/h08/WEG-MDW-brochure-50163906-pt.pdf
#   CWB  — WEG, "CWB - Contatores", 50042424, pág. 7
#          static.weg.net/medias/downloadcenter/hb6/h0a/WEG-contatores-CWB-50042424-pt.pdf
#
# Dois cuidados que mudam o resultado:
#   1. O valor do disjuntor é POR POLO. Um tripolar de 32 A dissipa 3 x 6 = 18 W,
#      não 6 W. Esquecer isso subdimensiona a ventilação por um fator de três.
#   2. É a máxima ADMITIDA pela norma, não a medida típica. Erra pro lado do
#      frio — que é o lado certo de errar em painel.
FONTE_MDWH = "WEG MDW 50163906 p.17 (05/08/2026)"
FONTE_CWB = "WEG CWB 50042424 p.7 (05/08/2026)"

BASE_DISSIPACAO = [
    ["Disjuntor", "MDWH", "Minidisjuntor In < 10 A — por polo", "WEG", "", 3, 3, FONTE_MDWH],
    ["Disjuntor", "MDWH", "Minidisjuntor 10 a 16 A — por polo", "WEG", "", 3.5, 3.5, FONTE_MDWH],
    ["Disjuntor", "MDWH", "Minidisjuntor 16 a 25 A — por polo", "WEG", "", 4.5, 4.5, FONTE_MDWH],
    ["Disjuntor", "MDWH", "Minidisjuntor 25 a 32 A — por polo", "WEG", "", 6, 6, FONTE_MDWH],
    ["Disjuntor", "MDWH", "Minidisjuntor 32 a 40 A — por polo", "WEG", "", 7.5, 7.5, FONTE_MDWH],
    ["Disjuntor", "MDWH", "Minidisjuntor 40 a 50 A — por polo", "WEG", "", 9, 9, FONTE_MDWH],
    ["Disjuntor", "MDWH", "Minidisjuntor 50 a 63 A — por polo", "WEG", "", 13, 13, FONTE_MDWH],
    ["Disjuntor", "MDWH", "Minidisjuntor 63 a 100 A — por polo", "WEG", "", 15, 15, FONTE_MDWH],
    ["Disjuntor", "MDWH", "Minidisjuntor 100 a 125 A — por polo", "WEG", "", 20, 20, FONTE_MDWH],
    ["Contator", "CWB", "Bobina CC, contator até 38 A (máximo)", "WEG", 5.8, "", 5.8, FONTE_CWB],
    ["Contator", "CWB", "Bobina CA, contator até 38 A — 7,5 VA", "WEG", 7.5, "", 7.5, FONTE_CWB],
    ["Fonte 24 VCC", "", "CALCULADO: Psaída x (1/rend - 1). Ex.: 240 W a 93% = 18 W", "", "", "", "", "rendimento do datasheet do modelo"],
    ["Inversor", "", "CALCULADO: Pmotor x (1 - rend). Ex.: 5,5 kW a 97% = 165 W", "", "", "", "", "rendimento do datasheet do modelo"],
]


def design_termico():
    wb, ws = novo("Dissipação")
    carimbo(ws, 7)
    fim = tabela(
        ws, 5,
        ["Instalação", "Tag", "Código", "Descrição", "Dissipação unit. (W)", "Qtd", "Dissipação total (W)"],
        [14, 12, 14, 40, 18, 8, 20],
        exemplos=[
            ["Painel 1", "", "", "", "", "", "=E6*F6"],
            ["Painel 1", "", "", "", "", "", "=E7*F7"],
            ["Painel 1", "", "", "", "", "", "=E8*F8"],
        ],
        mono_cols=(2, 3, 5, 6, 7),
    )
    ws.cell(row=fim + 1, column=6, value="TOTAL").font = f_cabec
    t = ws.cell(row=fim + 1, column=7, value="=SUM(G6:G500)")
    t.font = Font(name=MONO, size=10, bold=True, color=TINTA)
    t.border = b_cabec
    nota(ws, fim + 3, "O total daqui é o valor que vai no campo 'Potência Dissipada Calculada' da capa do projeto.")

    # --- cálculo
    wc = wb.create_sheet("Cálculo")
    wc["A1"] = "Cálculo térmico do painel"
    wc["A1"].font = f_secao

    entradas = [
        ("Altura (mm)", ""), ("Largura (mm)", ""), ("Profundidade (mm)", ""),
        ("Material da caixa", "Aço"),
        ("Temperatura interna admissível (°C)", 50),
        ("Temperatura ambiente máxima (°C)", 35),
        ("Potência dissipada total (W)", "=Dissipação!G10"),
    ]
    linha = 3
    for rotulo, valor in entradas:
        wc.cell(row=linha, column=1, value=rotulo).font = f_corpo
        c = wc.cell(row=linha, column=2, value=valor)
        c.font = f_mono
        c.border = b_linha
        c.fill = fill_zona
        linha += 1
    wc.column_dimensions["A"].width = 38
    wc.column_dimensions["B"].width = 16
    wc.column_dimensions["D"].width = 38
    wc.column_dimensions["E"].width = 16

    dvm = DataValidation(type="list", formula1='"Aço,Inox,Alumínio,Plástico"', allow_blank=True)
    wc.add_data_validation(dvm)
    dvm.add("B6")

    wc.cell(row=3, column=4, value="Superfície efetiva (m²)").font = f_corpo
    wc.cell(row=3, column=5, value="=2*((B3*B4)+(B3*B5)+(B4*B5))/1000000").font = f_mono
    wc.cell(row=4, column=4, value="ΔT admissível (°C)").font = f_corpo
    wc.cell(row=4, column=5, value="=B7-B8").font = f_mono
    wc.cell(row=5, column=4, value="Coef. do material (W/m²·°C)").font = f_corpo
    wc.cell(row=5, column=5, value='=IF(B6="Aço",5.5,IF(B6="Inox",5.5,IF(B6="Alumínio",12,3.5)))').font = f_mono
    wc.cell(row=6, column=4, value="Potência dissipável natural (W)").font = f_corpo
    wc.cell(row=6, column=5, value="=E3*E4*E5").font = f_mono
    wc.cell(row=7, column=4, value="Excedente a remover (W)").font = f_corpo
    wc.cell(row=7, column=5, value="=MAX(0,B9-E6)").font = f_mono
    wc.cell(row=8, column=4, value="Solução").font = f_cabec
    wc.cell(row=8, column=5, value='=IF(E7=0,"Ventilação natural suficiente","Requer ventilação forçada ou climatização")').font = f_mono

    nota(wc, 11, "Coeficientes de superfície conforme prática usual (IEC 60890). Conferir contra o catálogo do fabricante da caixa.")
    nota(wc, 12, "Preencher a dissipação total na célula B9 com o total da aba Dissipação.")

    # --- base de dissipação, a preencher
    wbase = wb.create_sheet("Base de dissipação")
    wbase["A1"] = "Base de dissipação por componente"
    wbase["A1"].font = f_secao
    linhas_base = list(BASE_DISSIPACAO)
    linhas_base.append(["", "", "", "", "", "", "=E%d+F%d" % (4 + len(BASE_DISSIPACAO),
                                                              4 + len(BASE_DISSIPACAO)), ""])
    fim = tabela(
        wbase, 3,
        ["Categoria", "Código", "Descrição", "Fabricante", "Dissip. bobina (W)", "Dissip. contatos (W)", "Total (W)", "Fonte do dado"],
        [18, 14, 38, 16, 18, 18, 14, 40],
        exemplos=linhas_base,
        mono_cols=(2, 5, 6, 7),
    )
    fim = nota(wbase, fim + 1, "As linhas acima vieram do catálogo do fabricante, com documento e data na última coluna. Conferir na revisão do catálogo em uso.")
    fim = nota(wbase, fim, "Disjuntor: o valor é POR POLO — multiplicar pelo número de polos. E é o MÁXIMO da norma, não o típico medido: erra a favor do frio.")
    fim = nota(wbase, fim, "Fonte e inversor não têm valor de tabela: a dissipação sai do rendimento. Ver as duas últimas linhas.")
    fim = nota(wbase, fim + 1, "Esta base cresce a cada projeto e é ativo próprio.")
    nota(wbase, fim, "Nunca copiar base de dados de terceiros — o dado é público no datasheet, a compilação é sua.")
    salvar(wb, "MODELO - Design Térmico.xlsx")


# ------------------------------------------------------------------ 4. tags

INSTRUCAO_IMPRESSAO = [
    ("Impressora", ""),
    ("Modelo de etiqueta", ""),
    ("Orientação", ""),
    ("Coluna a imprimir", ""),
    ("Observação", ""),
]


def aba_impressao(wb):
    ws = wb.create_sheet("Impressão")
    ws["A1"] = "Instruções de impressão"
    ws["A1"].font = f_secao
    linha = 3
    for rotulo, valor in INSTRUCAO_IMPRESSAO:
        ws.cell(row=linha, column=1, value=rotulo).font = f_corpo
        c = ws.cell(row=linha, column=2, value=valor)
        c.font = f_mono
        c.border = b_linha
        c.fill = fill_zona
        linha += 1
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 44
    nota(ws, linha + 1, "Esta aba é parte da entrega: é ela que faz o cliente imprimir sem precisar te ligar.")
    return ws


def planilha_tags(nome_arquivo, titulo_aba, cabecalhos, larguras, nota_texto):
    wb, ws = novo(titulo_aba)
    carimbo(ws, len(cabecalhos))
    fim = tabela(ws, 5, cabecalhos, larguras,
                 exemplos=[[""] * len(cabecalhos)],
                 mono_cols=tuple(range(1, len(cabecalhos) + 1)))
    nota(ws, fim + 1, nota_texto)
    aba_impressao(wb)
    salvar(wb, nome_arquivo)


def identificacoes():
    planilha_tags(
        "MODELO - Painel - TAGs Fios.xlsx", "TAGs Fios Painel",
        ["Ordem", "Régua", "Instalação", "Identificação do fio", "Potencial"],
        [10, 12, 16, 24, 16],
        "Fios de 24V e 0V não levam tag nem luva — são identificados pela cor.")

    planilha_tags(
        "MODELO - Painel - TAGs Bornes.xlsx", "TAGs Bornes",
        ["Ordem", "Régua", "Instalação", "Identificação do borne"],
        [10, 12, 16, 24],
        "Uma linha por borne, na ordem física da régua.")

    planilha_tags(
        "MODELO - Painel - TAGs Dispositivos.xlsx", "TAGs Dispositivos",
        ["Ordem", "Instalação", "Tag", "Descrição do dispositivo"],
        [10, 16, 14, 44],
        "Todo dispositivo do layout precisa de etiqueta.")

    planilha_tags(
        "MODELO - Painel - TAGs Rele.xlsx", "TAGs Relés",
        ["Ordem", "Instalação", "Tag", "Função"],
        [10, 16, 14, 44],
        "Identificação da base e do relé.")

    planilha_tags(
        "MODELO - Campo - TAGs Fios.xlsx", "TAGs Fios Campo",
        ["Ordem", "Cabo", "Instalação", "Identificação do fio", "Destino"],
        [10, 12, 16, 24, 28],
        "Fios dos cabos de campo, por cabo.")

    planilha_tags(
        "MODELO - Campo - TAGs Cabos.xlsx", "TAGs Cabos",
        ["Ordem", "Tag do cabo", "Descrição", "Origem", "Destino"],
        [10, 14, 40, 24, 24],
        "Uma etiqueta por ponta de cabo — são duas por cabo.")


# ------------------------------------------------------------------ 5. instalação

def lista_de_instalacao():
    wb, ws = novo("Base de cabos")
    carimbo(ws, 9)
    fim = tabela(
        ws, 5,
        ["Tag", "Descrição do cabo", "Código", "Especificação", "Vias", "Bitola (mm²)", "Compr. (m)", "Origem", "Destino"],
        [10, 38, 14, 18, 8, 12, 12, 24, 24],
        exemplos=[["W0", "", "", "", "", "", "", "", ""], ["W1", "", "", "", "", "", "", "", ""]],
        mono_cols=(1, 3, 5, 6, 7),
    )
    nota(ws, fim + 1, "É a fonte do diagrama de interconexão e da lista de material de campo.")

    # acompanhamento
    wa = wb.create_sheet("Acompanhamento")
    carimbo(wa, 10)
    fim = tabela(
        wa, 5,
        ["Tarefa", "Instalação", "Tag", "Tag do cabo", "Material", "Código", "Compr. (m)",
         "Horas estimadas", "% conclusão", "Horas concluídas"],
        [34, 14, 10, 12, 32, 14, 10, 14, 12, 16],
        exemplos=[
            ["Fixação do painel", "", "", "", "", "-", "-", "", "", "=H6*I6"],
            ["Fixação das eletrocalhas", "", "", "", "", "-", "-", "", "", "=H7*I7"],
            ["", "", "", "", "", "", "", "", "", "=H8*I8"],
        ],
        mono_cols=(3, 4, 6, 7, 8, 9, 10),
    )
    wa.cell(row=fim + 1, column=7, value="TOTAL").font = f_cabec
    for col, formula in ((8, "=SUM(H6:H500)"), (10, "=SUM(J6:J500)")):
        c = wa.cell(row=fim + 1, column=col, value=formula)
        c.font = Font(name=MONO, size=10, bold=True, color=TINTA)
        c.border = b_cabec
    dvp = DataValidation(type="decimal", operator="between", formula1=0, formula2=1, allow_blank=True)
    dvp.prompt = "Percentual de 0 a 1"
    wa.add_data_validation(dvp)
    dvp.add("I6:I500")
    nota(wa, fim + 3, "As horas estimadas transformam a entrega de lista de material em plano de instalação — é o que nenhum concorrente entrega.")

    # material de instalação
    wm = wb.create_sheet("Material de instalação")
    carimbo(wm, 6)
    fim = tabela(wm, 5,
                 ["Código", "Descrição", "Unidade", "Qtd calculada", "Arredondamento", "Qtd a comprar"],
                 [14, 44, 10, 14, 16, 14],
                 exemplos=[["", "", "", "", "", "=IF(E6=\"\",D6,CEILING(D6,E6))"]],
                 mono_cols=(1, 3, 4, 5, 6))
    nota(wm, fim + 1, "Terminal se compra de 100 em 100, luva de 10 em 10 — a coluna de arredondamento evita comprar quantidade que não existe.")

    # acessórios
    for titulo, regra in (
        ("Acessórios por ocorrência", "Entra uma vez por ocorrência do item principal — identificação, luva, prensa-cabo."),
        ("Acessórios por comprimento", "Entra proporcional ao metro — corrugado, clip, abraçadeira."),
        ("Acessórios por item", "Entra por item instalado — flange, acabamento de eletrocalha."),
    ):
        wx = wb.create_sheet(titulo)
        wx["A1"] = titulo
        wx["A1"].font = f_secao
        cabec = ["Código principal", "Descrição"]
        larg = [16, 40]
        for i in range(1, 5):
            cabec += [f"Acessório {i}", f"Multiplicador {i}"]
            larg += [16, 14]
        fim = tabela(wx, 3, cabec, larg, exemplos=[[""] * len(cabec)], mono_cols=tuple(range(1, len(cabec) + 1)))
        nota(wx, fim + 1, regra)

    wf = wb.create_sheet("Acessórios fixos")
    wf["A1"] = "Acessórios fixos por obra"
    wf["A1"].font = f_secao
    fim = tabela(wf, 3, ["Código", "Descrição", "Unidade", "Qtd"], [14, 44, 10, 10],
                 exemplos=[["", "", "", ""]], mono_cols=(1, 3, 4))
    nota(wf, fim + 1, "Parafuso, arruela, bucha — quantidade fixa, não depende do tamanho da obra.")

    wfo = wb.create_sheet("Fornecedores")
    wfo["A1"] = "Fornecedores"
    wfo["A1"].font = f_secao
    fim = tabela(wfo, 3, ["Código", "Descrição", "Fabricante", "Fornecedor", "Contato"],
                 [14, 40, 18, 26, 26], exemplos=[["", "", "", "", ""]], mono_cols=(1,))
    nota(wfo, fim + 1, "Base própria, cresce a cada projeto.")

    salvar(wb, "MODELO - Lista de Instalação.xlsx")


# ------------------------------------------------------------------ execução

if __name__ == "__main__":
    print("Gerando planilhas modelo:")
    lista_de_materiais()
    arquitetura_clp()
    design_termico()
    identificacoes()
    lista_de_instalacao()
    print(f"\nSaída: {SAIDA}")
